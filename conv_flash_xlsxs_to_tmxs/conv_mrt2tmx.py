#!/usr/bin/env python3

#  This file is part of cApps.
#
#  This script is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This script is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with cApps.  If not, see <https://www.gnu.org/licenses/>.

# ############# AUTHORSHIP INFO ###########################################

__author__ = "Manuel Souto Pico"
__copyright__ = "Copyright 2021, cApps/cApStAn"
__credits__ = ["Manuel Souto Pico"]
__license__ = "GPL"
__version__ = "0.2.0"
__maintainer__ = "Manuel Souto Pico"
__email__ = "manuel.souto@capstan.be"
__status__ = "Testing / pre-production" # "Production"

# ############# IMPORTS ###########################################

import os
import sys
import re
import argparse
import xlrd
from yattag import Doc, indent
import pandas as pd
from pandas_ods_reader import read_ods
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#from openpyxl.utils.cell import column_index_from_string
from pprint import pprint
# import xml.dom.minidom
import collections


# ############# PROGRAM DESCRIPTION ###########################################

text = "This is *TM Workbook Converter*: it takes an IPSOS' MRT file where each \
column contains a language version and produces as many TMX files as target \
languages the workbook has. It assumes ENU as source language. Entries in the \
TMX files will have no particular order, but they contain no duplicates."

# intialize arg parser with a description
parser = argparse.ArgumentParser(description=text)
parser.add_argument("-V", "--version", help="show program version",
                    action="store_true")
parser.add_argument("-i", "--input", help="specify path to input file")
parser.add_argument("-m", "--mapping", help="specify path to mapping file, to know correspondence between different language codes")
parser.add_argument("-s", "--survey", help="specify the survey name that will be used to name other things")


# read arguments from the command line
args = parser.parse_args()

# check for -V or --version
if args.version:
    print("This is Flash Prepp/Extract utility version 0.1.0")
    sys.exit()

if args.input and args.mapping and args.survey:
    path_to_mrt_file = args.input.rstrip('/')
    path_to_mapping_file = args.mapping.rstrip('/')
    survey = args.survey.strip()
else:
    print("Arguments -i or -m or -s not found.")
    sys.exit()


## no config for now, constants are few and hard-coded
config = dict()
config['source_lang'] = 'ENU'
config['header_row'] = 1
config['sheet_name'] = 'MDD Labels'
config['container'] = 'Eurobarometer'
config['tmx_file_name'] = '<container>, <survey>, <target_lang>, MRT'
print(config)


# ############# FUNCTIONS #####################################################

# def map_langtag_loc(df, tag, xfrom, xto): # same output as map_langtag
#     # print(langtags.loc[langtags.cApStAn == x, 'OmegaT'].values[0])
#     return df.loc[df[xfrom] == tag, xto].values[0]


def map_langtag(df, tag, xfrom, xto): # same output as map_langtag_loc
    # langtags_dict = dict(zip(langtags['cApStAn'], langtags['OmegaT']))
    langtags_dict = dict(zip(df[xfrom], df[xto]))
    return langtags_dict[tag]


# def get_config(wb, sheet_idx):
#     sheet = wb.sheet_by_index(sheet_idx)
#     parameters = sheet.col_values(0)
#     values = sheet.col_values(1)
#     return dict(zip(parameters, values))


# def get_data(wb, sheet_idx, source_col, target_col):
#     # if sheet0 has name "config", if not use sheet0
#     sheet = wb.sheet_by_index(sheet_idx)
#     # print(sheet.cell_value(2,1))
#     # print(sheet.nrows)
#     # print(sheet.ncols)
#     # print(sheet.row_values(0))
#     source_texts = sheet.col_values(source_col)
#     target_texts = sheet.col_values(target_col)
#     return set(zip(source_texts, target_texts))

def get_langpair_from_sheet(sheet, source_col, target_col):
    source_texts = [cell.value for cell in sheet[source_col]]
    target_texts = [cell.value for cell in sheet[target_col]]
    return [tu for tu in list(zip(source_texts, target_texts))  if tu[0] != None and tu[1] != None]


def get_langpair_set_from_sheet(sheet, source_col, target_col):
    source_texts = [cell.value for cell in sheet[source_col]]
    target_texts = [cell.value for cell in sheet[target_col]]
    return [tu for tu in set(zip(source_texts, target_texts)) if tu[0] != None and tu[1] != None]


# def get_headers(wb, sheet_idx, row_idx):
#     sheet = wb.sheet_by_index(sheet_idx)
#     return sheet.row_values(row_idx)


def build_tmx(langpair_set, xml_source_lang, xml_target_lang):
    # convert to tmx
    doc, tag, text = Doc().tagtext()

    doc.asis('<?xml version="1.0" encoding="UTF-8"?>')
    with tag('tmx', version="1.4"):
        with tag('header', creationtool="cApps", creationtoolversion="2020.10",
                 segtype="paragraph", adminlang="en",
                 datatype="HTML", srclang=xml_source_lang):
            doc.attr(
                ('o-tmf', "omt") # o_tmf="omt",
            )
            text('')
        with tag('body'):
            for tu in langpair_set:
                src_txt = str(tu[0]).strip()
                tgt_txt = str(tu[1]).strip()
                with tag('tu'):
                    with tag('tuv'):
                        doc.attr(
                            ('xml:lang', xml_source_lang)
                        )
                        with tag('seg'):
                            text(src_txt)
                    with tag('tuv'):
                        doc.attr(
                            ('xml:lang', xml_target_lang)
                        )
                        with tag('seg'):
                            text(tgt_txt)

    tmx_output = indent(
        doc.getvalue(),
        indentation=' '*2,
        newline='\r\n'
    )
    return tmx_output  # .replace("o_tmf=", "o-tmf=")


# def get_langs(wb, config): # not used for mrt
#     return [x for x in get_headers(wb, 1, int(config['header_row']))
#             if re.match(r'[a-z]{3}-[A-Z]{3}', x) and x != config['source_lang']]


def write_tmx_file(config, tmx_output):
    # build filename
    config['tmx_file_name'] = config['tmx_file_name'].replace('<', '').replace('>', '')
    fn_parts = [config[x.strip()] if x.strip() in config.keys()
                else x.strip()
                for x in config['tmx_file_name'].split(',')]

    # writing output
    filename = "_".join(fn_parts) + ".tmx"
    print("Writing TMX output to file " + filename)
    with open("output/" + filename, "w") as f:
        print(tmx_output, file=f) # used


def convert_mrt_to_tmx_files(path_to_file, langtags_csv):

    sheet_name = config['sheet_name']

    # extract langtags df
    #langtag_mapping_df = read_ods(path_to_mapping_file, 1) # sheet_idx
    langtags_df = pd.read_excel(path_to_mapping_file, sheet_name='mapping')

    # extract MRT
    wb = load_workbook(path_to_mrt_file, data_only=True)
    sheet = wb[sheet_name]
    sheet = wb.active
    first_row = sheet[1] # 1 = config['header_row']
    #headers = [cell.value for cell in first_row]

    cols_data = [
        {
            'col_letter': get_column_letter(cell.column),
            'col_name': cell.value,
            'mrt_code': re.search('([A-Z]{3})\s.+', cell.value).group(1)
        }
        for cell in first_row
        if re.search('([A-Z]{3})\s.+', cell.value) != None]

    # add omegat tag and capstan code (more readable to add them separately than in the list comprehenion above)
    for item in cols_data:
        item['omegat_tag'] = map_langtag(langtags_df, item['mrt_code'], xfrom = 'MRT', xto = 'XML')
        item['capstan_code'] = map_langtag(langtags_df, item['mrt_code'], xfrom = 'MRT', xto = 'cApStAn')
    # see below @1 other alternative ways of doing this


    # transform list to dictionary with mrt code as key for each
    langs_data = {item['mrt_code']: item for item in cols_data}

    for mrt_code, version in langs_data.items():
        if mrt_code == config['source_lang']:
            continue

        mrt_srclang = config['source_lang']
        srclang = langs_data[mrt_srclang]
        langpair_set = get_langpair_set_from_sheet(sheet, source_col=srclang['col_letter'], target_col=version['col_letter'])
        tmx_output = build_tmx(langpair_set, srclang['omegat_tag'], version['omegat_tag'])
        lang_config = dict(config, target_lang=version['capstan_code'], survey=survey)
        write_tmx_file(lang_config, tmx_output)



# ############# EXECUTION #####################################################

print(path_to_mrt_file)
print(path_to_mapping_file)

convert_mrt_to_tmx_files(path_to_mrt_file, path_to_mapping_file)


# TODO
# split by linebreak
# segment  by punctuation ???
# extract to functions
sys.exit()
# --------------------------------------------------------------------------------------------------------------------------------

# move the output files to /Volumes/data/company/IPSOS/EUROBAROMETER_FLASH_2.0/09_ASSETS/01_Incoming/from_MRT_source

#@1
#langsets = [item['omegat_tag'] = map_langtag(langtags_df, item['mrt_code'], xfrom = 'MRT', xto = 'XML') for item in cols_data]
#langsets = [item.update({'c':3,'d':4}) for item in cols_data]
#langsets = list(map(lambda item: dict(item, omegat_tag= map_langtag(langtags_df, item['mrt_code'], xfrom = 'MRT', xto = 'XML')     ), cols_data))
'''
# with list comprehension, more functional, less readable
langsets = [dict(item,
        omegat_tag= map_langtag(langtags_df, item['mrt_code'], xfrom = 'MRT', xto = 'XML'),
        capstan_code= map_langtag(langtags_df, item['mrt_code'], xfrom = 'MRT', xto = 'cApStAn')    )
    for item in cols_data]
'''

languages = list(filter(lambda h: re.search('([A-Z]{3})\s.+', h), headers))
language_codes = [re.search('([A-Z]{3})\s.+', h).group(1) for h in headers if re.search('([A-Z]{3})\s.+', h) != None]
code_to_language_map = dict(zip(language_codes, languages))


# ############# WIP ###########################################################

# done:
# 1. parse excel, extract each language pair (cols 1.3)
# 2. put each language pair in one tmx

# todo:
# 1. segment paragraphs whenever possible
# 2. clean up tags
