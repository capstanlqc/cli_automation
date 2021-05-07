# cd <parent_folder>
# pipenv shell --three
# pipenv install pandas_ods_reader openpyxl yattag

import os
import os.path
import re
from yattag import Doc, indent
from pandas_ods_reader import read_ods
import pandas as pd
from openpyxl import load_workbook

# global constants
# get from config file # <--- arg to the program
source_langtag = "en-GB" 
langtag_mapping_path = './language_codes_eb-flash.ods'
path_to_files = './flash_files' # preferable an absolute path
files = os.listdir(path_to_files)
fl419_filename_template = 'FL419 <lang>.xls' # read from config
fl419_filename_pattern = re.compile(r'FL419\s+([A-Z]+)\.xls')
sheet_name_to_extract = 'TRANSLATION'
config_df = read_ods('./config.ods', 1) # sheet_idx
langtag_mapping_df = read_ods(langtag_mapping_path, 1) # sheet_idx


def get_config(df):
    #sheet = wb.sheet_by_index(sheet_idx)
    parameters = df.iloc[:, 0] # first column
    values = df.iloc[:, 1] # second column
    return dict(zip(parameters, values))


# def flash_version(file):
#   name has pattern && extract locale from file_name
#     if locale in locale_list
#       return get bcp47_tag

# for file in files:
#   if flash_version(file): # returns bcp47_tag
#       data = extract_data(file)  # data is dict where { coord: data }
#       multidata[locale] = data
#
#       for version in multidata:
#           if not source:
#               make_tmx(source, version)

def extract_version(path_to_xls_file, sheet_name_to_extract):
    # openpyxl does not support the old .xls file format 
    """ openpyxl_wb = load_workbook(path_to_xls_file)
        sheet = openpyxl_wb.get_sheet_by_name(sheet_name_to_extract_to_extract)
        for row in sheet.iter_rows():
            for cell in row:
                print(cell.coordinate)
                print(cell.value) """
    
    if os.path.isfile(path_to_xls_file): # file exists
        xls = pd.ExcelFile(path_to_xls_file)
        df = pd.read_excel(xls, sheet_name_to_extract, header=None)
        series = df.unstack()
        series.index = ['{}_{}'.format(x, y) for x, y in series.index ]
        return series.to_dict()
    else:
        print(f"File '{path_to_xls_file}' does not exist")


def map_langtag(tag, xfrom, xto):
    # langtags_dict = dict(zip(langtags['cApStAn'], langtags['OmegaT']))
    langtags_dict = dict(zip(langtag_mapping_df[xfrom], langtag_mapping_df[xto]))
    return langtags_dict[tag]


def source_file():
    source_kantar_langtag = map_langtag(source_langtag, 'XML', 'Kantar')
    filename = fl419_filename_template.replace('<lang>', source_kantar_langtag)
    #filename = f'FL419 {source_kantar_langtag}.xls'
    if filename in files:
        return filename
    else:
        False


def get_langtag_of_file(file):
    m = re.search(fl419_filename_pattern, file)
    if m:
        kantar_langtag = m.group(1)
        print(kantar_langtag)
        return map_langtag(kantar_langtag, 'Kantar', 'XML')


def build_tmx(langpair_dict, xml_source_lang, xml_target_lang):
    # convert to tmx
    doc, tag, text = Doc().tagtext()
    langpair_set = langpair_dict.items()

    doc.asis('<?xml version="1.0" encoding="UTF-8"?>')
    with tag('tmx', version="1.4"):
        with tag('header', creationtool="cApps", creationtoolversion="2021.02",
                 segtype="paragraph", adminlang="en",
                 datatype="HTML", srclang=xml_source_lang):
            doc.attr(
                ('o-tmf', "omt") # o_tmf="omt",
            )
            text('')
        with tag('body'):
            for tu in langpair_set:
                
                src_txt = str(tu[0]).strip()
                tgt_txt = str(tu[1]).strip()

                if src_txt != 'nan' and \
                    tgt_txt != 'nan' and \
                    src_txt != tgt_txt:
                    
                    with tag('tu'):
                        with tag('tuv'):
                            doc.attr(
                                ('xml:lang', xml_source_lang)
                            )
                            with tag('seg'):
                                text(src_txt)
                        with tag('tuv'):
                            doc.attr(
                                ('xml:lang', xml_target_lang)
                            )
                            with tag('seg'):
                                text(tgt_txt)

    tmx_output = indent(
        doc.getvalue(),
        indentation=' '*2,
        newline='\r\n'
    )
    return tmx_output  # .replace("o_tmf=", "o-tmf=")


def write_tmx_file(config, tmx_output):
    # build filename
    config['tmx_file_names'] = config['tmx_file_names'].replace('<', '').replace('>', '')
    fn_parts = [config[x.strip()] if x.strip() in config.keys()
                else x.strip()
                for x in config['tmx_file_names'].split(',')]

    # writing output
    filename = "_".join(fn_parts) + ".tmx"
    print("Writing TMX output to file " + filename)
    with open("output/" + filename, "w") as f:
        print(tmx_output, file=f)



def traverse_dir():

    config = get_config(config_df)
    
    print(source_langtag)
    # check whether there is a source version before anything
    if source_file():
        print(f"Source language ({source_file()}) version file found") # log
        path_to_file = os.path.join(path_to_files, source_file())
        source_text = extract_version(path_to_file, sheet_name_to_extract)
    else:
        print(f"Source language ({source_langtag}) version file NOT found") # log
        return

    for filename in files:
        xml_tag = get_langtag_of_file(filename)
        path_to_file = os.path.join(path_to_files, filename)

        if xml_tag:
            if xml_tag != source_langtag:
                print("I'll try to create TM for " + xml_tag) # log
                target_text = extract_version(path_to_file, sheet_name_to_extract)

                langpair = {source_text[key]: target_text[key] 
                    for key in source_text.keys() & target_text.keys()}

                tmx_output = build_tmx(langpair, source_langtag, xml_tag)
                lang_config = dict(config, target_lang=xml_tag)
                write_tmx_file(lang_config, tmx_output)
        else:
            str = f"Unable to create TM for file '{filename}', unknown correspondence in BCP47."
            print(str) # log



def main_function():

    corres = dict(zip(langtag_mapping_df.Kantar, langtag_mapping_df.XML))

    print(langtag_mapping_df.Kantar[7])
    print(corres)

    xml_tag = map_langtag('EEE', 'Kantar', 'XML')
    print(xml_tag)

    traverse_dir()


# execution
main_function()