# cd <parent_folder>
# pipenv shell --three
# pipenv install pandas_ods_reader openpyxl yattag pylint argparse

#  This file is part of cApps.
#
#  This script is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This script is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with cApps.  If not, see <https://www.gnu.org/licenses/>.

# ############# AUTHORSHIP INFO ###########################################

__author__ = "Manuel Souto Pico"
__copyright__ = "Copyright 2021, cApps/cApStAn"
__credits__ = ["Manuel Souto Pico"]
__license__ = "GPL"
__version__ = "0.1.0"
__maintainer__ = "Manuel Souto Pico"
__email__ = "manuel.souto@capstan.be"
__status__ = "Testing / pre-production" # "Production"

# ############# IMPORTS ###########################################

import time
import logging
import sys
import os
sys.path.append(os.path.abspath('../py_mods'))
from func import my_func
import os.path
import argparse
import re
from yattag import Doc, indent #pylint: disable=E401
from pandas_ods_reader import read_ods #pylint: disable=E401
import pandas as pd #pylint: disable=E401
from openpyxl import load_workbook #pylint: disable=E401
import datetime
import pprint
pp = pprint.PrettyPrinter(indent=4)

my_func("my friend")


# ############# PROGRAM DESCRIPTION ###########################################

text = "This program converts pairs of FL419 spreadsheets into translation memories."

# intialize arg parser with a description
parser = argparse.ArgumentParser(description=text)
parser.add_argument("-V", "--version", help="show program version",
					action="store_true")
parser.add_argument("-c", "--config", help="specify path to config file, by default at ./config/config.ods")

# read arguments from the command line
args = parser.parse_args()

# check for -V or --version
if args.version:
	print("This is Flash TM maker version 0.1.0")
	sys.exit()
if args.config:
	print("Using preferences from %s" % args.config)

# ############# GLOBAL CONSTANTS ###########################################

source_lang = "en" # default
# global constants
# get from config file # <--- arg to the program
container = "Generic"
source_lang = "en-GB"
langtag_mapping_path = './config/ipsos_language_codes.xlsx'
path_to_files = './flash_files' # preferable an absolute path
files = os.listdir(path_to_files)
fl419_filename_template = 'FL419 <lang>.xls' # read from config

fl419_filename_pattern = re.compile(r'FL419\s+([A-Z]+)\.xls')
name_of_sheet_to_extract = 'TRANSLATION'
config_df = read_ods('./config/config.ods', 1) # sheet_idx
#langtag_mapping_df = read_ods(langtag_mapping_path, 1) # sheet_idx
langtag_mapping_df = pd.read_excel(langtag_mapping_path, 0) # sheet_idx



ts = time.gmtime()
formatted_ts = time.strftime("%Y%m%d_%H%M%S", ts)
logfile = '_log/' + formatted_ts + '.log'
#open(log_fname, "a")
print(f"The log will be printed to file '{logfile}'")


logging.basicConfig(filename=logfile, level=logging.DEBUG) # encoding='utf-8', 
# https://docs.python.org/3/howto/logging.html#logging-basic-tutorial

#sys.exit()


def get_config(df):
	#sheet = wb.sheet_by_index(sheet_idx)
	parameters = df.iloc[:, 0] # first column
	values = df.iloc[:, 1] # second column
	return dict(zip(parameters, values))


# def flash_version(file):
#   name has pattern && extract locale from file_name
#	 if locale in locale_list
#	   return get bcp47_tag

# for file in files:
#   if flash_version(file): # returns bcp47_tag
#	   data = extract_data(file)  # data is dict where { coord: data }
#	   multidata[locale] = data
#
#	   for version in multidata:
#		   if not source:
#			   make_tmx(source, version)

def extract_version(path_to_xls_file, name_of_sheet_to_extract):
	# openpyxl does not support the old .xls file format 
	""" openpyxl_wb = load_workbook(path_to_xls_file)
		sheet = openpyxl_wb.get_sheet_by_name(name_of_sheet_to_extract_to_extract)
		for row in sheet.iter_rows():
			for cell in row:
				print(cell.coordinate)
				print(cell.value) """
	
	if os.path.isfile(path_to_xls_file): # file exists
		xls = pd.ExcelFile(path_to_xls_file)
		df = pd.read_excel(xls, name_of_sheet_to_extract, header=None)
		series = df.unstack()
		series.index = ['{}_{}'.format(x, y) for x, y in series.index ]
		return series.to_dict()
	else:
		logging.warning(f" File '{path_to_xls_file}' does not exist")


def map_langtag(tag, xfrom, xto):
	# langtags_dict = dict(zip(langtags['cApStAn'], langtags['OmegaT']))
	langtags_dict = dict(zip(langtag_mapping_df[xfrom], langtag_mapping_df[xto]))
	#langtags_dict_clean = dict(filter(lambda tag: isinstance(tag, str), langtags_dict.items() ))
	clean_langtags_dict = {k: v for k, v in langtags_dict.items() if isinstance(k, str) and isinstance(v, str) }

	if tag in clean_langtags_dict: # tag is a key in dict
		return clean_langtags_dict[tag]
	else:
		return False


def source_file():
	source_kantar_langtag = map_langtag(source_lang, 'XML', 'Kantar')
	if not source_kantar_langtag:
		return False
	filename = fl419_filename_template.replace('<lang>', source_kantar_langtag)
	#filename = f'FL419 {source_kantar_langtag}.xls'
	if filename in files:
		return filename
	else:
		return False


def get_langtag_of_file(file):

	match = re.search(config["fl419_filename_pattern"], file)
	if match:
		kantar_langtag = match.group(1)
		logging.info(f" Kantar's langtag: '{kantar_langtag}'")
		return map_langtag(kantar_langtag, 'Kantar', 'XML')
	else:
		return False


def build_tmx(langpair_dict, xml_source_lang, xml_target_lang):
	# convert to tmx
	doc, tag, text = Doc().tagtext()
	langpair_set = langpair_dict.items()

	doc.asis('<?xml version="1.0" encoding="UTF-8"?>')
	with tag('tmx', version="1.4"):
		with tag('header', creationtool="cApps", creationtoolversion="2021.02",
				 segtype="paragraph", adminlang="en",
				 datatype="HTML", srclang=xml_source_lang):
			doc.attr(
				('o-tmf', "omt") # o_tmf="omt",
			)
			text('')
		with tag('body'):
			for tu in langpair_set:
				
				src_txt = str(tu[0]).strip()
				tgt_txt = str(tu[1]).strip()

				if src_txt != 'nan' and \
					tgt_txt != 'nan' and \
					src_txt != tgt_txt:
					
					with tag('tu'):
						with tag('tuv'):
							doc.attr(
								('xml:lang', xml_source_lang)
							)
							with tag('seg'):
								text(src_txt)
						with tag('tuv'):
							doc.attr(
								('xml:lang', xml_target_lang)
							)
							with tag('seg'):
								text(tgt_txt)

	tmx_output = indent(
		doc.getvalue(),
		indentation=' '*2,
		newline='\r\n'
	)
	return tmx_output  # .replace("o_tmf=", "o-tmf=")


def write_tmx_file(lang_config, tmx_output):

	# build filename
	lang_config['tmx_file_name'] = lang_config['tmx_file_name'].replace('<', '').replace('>', '')
	fn_parts = [lang_config[x.strip()] if x.strip() in lang_config.keys()
				else x.strip()
				for x in lang_config['tmx_file_name'].split(',')]

	# writing output
	filename = "_".join(fn_parts) + ".tmx" # <- this includes the XML language tag
	# replace XML tag in filename with cApStAn code
	filename = filename.replace(lang_config['target_lang'], lang_config['capstan_code'])
	logging.info(f" Writing TMX output to file '{filename}'")
	with open("output/" + filename, "w") as f:
		print(tmx_output, file=f)


def process_files():

	global config
	config = get_config(config_df)

	# check whether there is a source version before anything
	if source_file():
		logging.info(f" Source language version file ('{source_file()}') found")
		path_to_file = os.path.join(path_to_files, source_file())
		source_text = extract_version(path_to_file, name_of_sheet_to_extract)
	else:
		logging.warning(f" Source language ({source_lang}) version file NOT found") 
		return

	for filename in files:
		# xml_tag to be used internally in the TMX file (as XML tag)
		xml_tag = get_langtag_of_file(filename)
		# capstan_tag to be used in the filename
		capstan_code = map_langtag(xml_tag, 'XML', 'cApStAn')
		path_to_file = os.path.join(path_to_files, filename)

		if xml_tag:
			if xml_tag != source_lang: # not nan
				#xml_tag_str = str(xml_tag)
				logging.info(f" Trying to create TM for '{xml_tag}'") 
				target_text = extract_version(path_to_file, name_of_sheet_to_extract)

				langpair = {source_text[key]: target_text[key] 
					for key in source_text.keys() & target_text.keys()}

				tmx_output = build_tmx(langpair, source_lang, xml_tag)
				lang_config = dict(config, target_lang=xml_tag, capstan_code=capstan_code)
				write_tmx_file(lang_config, tmx_output)
		else:
			str = f" Unable to create TM for file '{filename}', unknown correspondence in BCP47."
			logging.warning(str) # log


if __name__ == "__main__":

	begin_time = datetime.datetime.now()
	logging.info(f' Started at {begin_time}')

	corres = dict(zip(langtag_mapping_df.Kantar, langtag_mapping_df.XML))
	corres_clean = {k: v for k, v in corres.items() if isinstance(k, str) and isinstance(v, str) }

	process_files()
	finish_time = datetime.datetime.now()
	logging.info(f' Finished at {finish_time}')
	logging.info(f' It look {finish_time - begin_time} to run the script')


"""
TODO
- make sure --config works (parameter overrides default)
- get constants from config file
- create clean df's once 
- put common functions in a module
"""
